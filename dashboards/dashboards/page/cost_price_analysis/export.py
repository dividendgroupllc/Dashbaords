"""Выгрузка таблицы себестоимости в .xlsx.

Книга повторяет экран: лист «Себестоимость» — числовые значения с той же
красно-зелёной заливкой, лист «Изменение %» — проценты к предыдущему месяцу.
Значения пишутся числами (не текстом), чтобы в Excel по ним сразу считались
формулы, сводные таблицы и графики.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dashboards.dashboards.dashboard_data import MONTH_LABELS

HEADER_FILL = PatternFill("solid", fgColor="F2F6FC")
UP_FILL = PatternFill("solid", fgColor="FDECEB")
DOWN_FILL = PatternFill("solid", fgColor="E7F6EF")

HEADER_FONT = Font(name="Calibri", bold=True, color="33496A")
UP_FONT = Font(name="Calibri", bold=True, color="C4161C")
DOWN_FONT = Font(name="Calibri", bold=True, color="0F8A5F")
VALUE_FONT = Font(name="Calibri", bold=True, color="1F3050")

MONEY_FORMAT = "#,##0.00"
PERCENT_FORMAT = "+0.0%;-0.0%;0.0%"

LEAD_COLUMNS = ("Товар", "Группа", "Ед.")
LEAD_WIDTHS = (44, 22, 9)
MONTH_WIDTH = 14


def build_workbook(context: dict[str, Any], rows: list[dict[str, Any]]) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    _write_sheet(workbook, "Себестоимость", context, rows, mode="value")
    _write_sheet(workbook, "Изменение %", context, rows, mode="change")

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _write_sheet(
    workbook: openpyxl.Workbook,
    title: str,
    context: dict[str, Any],
    rows: list[dict[str, Any]],
    mode: str,
) -> None:
    sheet = workbook.create_sheet(title)
    months = context.get("months") or []

    header = list(LEAD_COLUMNS) + [
        MONTH_LABELS[int(month["index"]) - 1] for month in months
    ]
    sheet.append(header)
    for column_index in range(1, len(header) + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center" if column_index > len(LEAD_COLUMNS) else "left")

    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=row.get("item_name"))
        sheet.cell(row=row_index, column=2, value=row.get("item_group"))
        sheet.cell(row=row_index, column=3, value=row.get("uom"))

        for month_index, cell_data in enumerate(row.get("cells") or []):
            column_index = len(LEAD_COLUMNS) + month_index + 1
            cell = sheet.cell(row=row_index, column=column_index)
            if not cell_data:
                continue

            direction = cell_data.get("direction")
            if mode == "value":
                cell.value = round(float(cell_data.get("value") or 0), 2)
                cell.number_format = MONEY_FORMAT
                # Как на экране: месяцы без изменения выглядят одинаково,
                # выделяются только рост и снижение.
                cell.font = VALUE_FONT
            else:
                change = cell_data.get("change")
                if change is None or direction not in ("up", "down"):
                    continue
                # Доля, а не «проценты числом»: Excel сам покажет 2,5 % и посчитает по ней.
                cell.value = round(float(change) / 100, 6)
                cell.number_format = PERCENT_FORMAT

            if direction == "up":
                cell.fill = UP_FILL
                cell.font = UP_FONT
            elif direction == "down":
                cell.fill = DOWN_FILL
                cell.font = DOWN_FONT

    for column_index, width in enumerate(LEAD_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for month_index in range(len(months)):
        column_letter = get_column_letter(len(LEAD_COLUMNS) + month_index + 1)
        sheet.column_dimensions[column_letter].width = MONTH_WIDTH

    # Шапка и колонки товара остаются на месте при прокрутке.
    sheet.freeze_panes = f"{get_column_letter(len(LEAD_COLUMNS) + 1)}2"
    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"


def build_filename(context: dict[str, Any]) -> str:
    return f"sebestoimost-{context.get('selected_year') or ''}"
